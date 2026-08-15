"""
Validasi silang: bandingkan zscore_who_permenkes_2020.csv (hasil ekstraksi PDF Permenkes)
terhadap tabel Z-score resmi WHO dalam format Excel (folder who_excel_official/).

Tujuan: memastikan proses ekstraksi PDF tidak menghasilkan kesalahan pembacaan angka,
dengan membandingkan setiap baris terhadap sumber data primer resmi WHO.

Jalankan (butuh file .xlsx di who_excel_official/, lihat README.md cara unduhnya):
    python crosscheck_who.py
"""

import csv
import openpyxl
from pathlib import Path

BASE_DIR = Path(__file__).parent
CSV_PATH = BASE_DIR / "zscore_who_permenkes_2020.csv"
WHO_DIR = BASE_DIR / "who_excel_official"
TOLERANSI = 0.05

# load CSV hasil ekstraksi PDF Permenkes jadi dict lookup
permenkes = {}
with open(CSV_PATH, newline="", encoding="utf-8") as f:
    for row in csv.DictReader(f):
        key = (row["indeks"], row["jenis_kelamin"], round(float(row["nilai_x"]), 2))
        permenkes[key] = [
            float(row["sd3neg"]), float(row["sd2neg"]), float(row["sd1neg"]),
            float(row["median"]), float(row["sd1pos"]), float(row["sd2pos"]), float(row["sd3pos"]),
        ]

# daftar file WHO -> (indeks, gender)
mapping = [
    ("wfa_boys_zscores.xlsx", "BB/U", "laki-laki"),
    ("wfa_girls_zscores.xlsx", "BB/U", "perempuan"),
    ("lhfa_boys_0-2_zscores.xlsx", "PB/U", "laki-laki"),
    ("lhfa_girls_0-2_zscores.xlsx", "PB/U", "perempuan"),
    ("lhfa_boys_2-5_zscores.xlsx", "TB/U", "laki-laki"),
    ("lhfa_girls_2-5_zscores.xlsx", "TB/U", "perempuan"),
    ("wfl_boys_zscores.xlsx", "BB/PB", "laki-laki"),
    ("wfl_girls_zscores.xlsx", "BB/PB", "perempuan"),
    ("wfh_boys_zscores.xlsx", "BB/TB", "laki-laki"),
    ("wfh_girls_zscores.xlsx", "BB/TB", "perempuan"),
]

total_baris = 0
total_cocok = 0
total_gak_ada = 0
selisih_maks_global = 0
detail_gagal = []

for fname, indeks, gender in mapping:
    wb = openpyxl.load_workbook(WHO_DIR / fname, data_only=True)
    ws = wb.active
    header = [str(c.value).strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx_x = 0  # kolom pertama = Month/Length/Height
    idx_sd_start = header.index("SD3neg")

    baris_ini = 0
    cocok_ini = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        x = round(float(row[idx_x]), 2)
        who_vals = [float(v) for v in row[idx_sd_start:idx_sd_start + 7]]
        key = (indeks, gender, x)

        if key not in permenkes:
            total_gak_ada += 1
            continue

        p_vals = permenkes[key]
        baris_ini += 1
        total_baris += 1

        selisih = [abs(a - b) for a, b in zip(who_vals, p_vals)]
        selisih_maks = max(selisih)
        selisih_maks_global = max(selisih_maks_global, selisih_maks)

        if selisih_maks <= TOLERANSI:
            cocok_ini += 1
            total_cocok += 1
        else:
            detail_gagal.append((indeks, gender, x, who_vals, p_vals, selisih_maks))

    print(f"{fname:30s} | {indeks:6s} {gender:10s} | {cocok_ini}/{baris_ini} baris cocok (toleransi {TOLERANSI})")

print()
print("=" * 70)
print(f"TOTAL: {total_cocok}/{total_baris} baris cocok ({total_cocok/total_baris*100:.2f}%)")
print(f"Baris WHO yang gak ada padanan umur/panjang di CSV Permenkes: {total_gak_ada}")
print(f"Selisih maksimum yang pernah ditemukan: {selisih_maks_global}")

if detail_gagal:
    print(f"\nContoh baris tidak cocok ({len(detail_gagal)} total):")
    for indeks, gender, x, who_vals, p_vals, selisih in detail_gagal[:10]:
        print(f"  {indeks} {gender} x={x} | WHO={who_vals} | Permenkes={p_vals} | selisih_maks={selisih}")
